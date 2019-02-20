# -*- coding: utf-8 -*-
# (c) 2015 AvanzOSC
# License AGPL-3 - See http://www.gnu.org/licenses/agpl-3.0.html

from odoo import fields, models, exceptions, api, _
from io import BytesIO
import openpyxl
import base64

class ImportInventory(models.TransientModel):
    _name = 'import.inventory'
    _description = 'Import inventory'


    data = fields.Binary('File', required=True)
    name = fields.Char('Filename')

    @api.multi
    def action_import(self):
        """Load Inventory data from the CSV file."""
        inventory_obj = self.env['stock.inventory']
        inv_imporline_obj = self.env['stock.inventory.line']
        product_obj = self.env['product.product']
        inventory = inventory_obj
        fdata = self.data and base64.decodestring(self.data) or False
        file_input = BytesIO(fdata)
        file_input.seek(0)
        wb = openpyxl.load_workbook(file_input)
        sheet = wb._sheets[0]
        lot_id=False
        inventory =inventory_obj.browse(self._context.get('active_id'))
        for rowNum in range(2, sheet.max_row+1):
                if sheet.cell(row=rowNum, column=1).value:
                    product_code = sheet.cell(row=rowNum, column=1).value
                else:
                    continue
                product = product_obj.search([('default_code', '=', product_code)],limit=1)
                if sheet.max_column >= 2:
                    if float(sheet.cell(row=rowNum, column=2).value):
                        product_qty = float(sheet.cell(row=rowNum, column=2).value)
                if sheet.max_column >= 3:
                    lot_id = str(sheet.cell(row=rowNum, column=3).value)
                inv_imporline_obj.create({
                    'product_id': product.id,
                    'product_qty' : product_qty,
                    'product_uom_id' : product.uom_id.id,
                    'prod_lot_id' : lot_id or False,
                    'location_id': inventory.location_id.id,
                    'inventory_id':inventory.id,
                })


